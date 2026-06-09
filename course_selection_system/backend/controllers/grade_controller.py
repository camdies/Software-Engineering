"""
backend/controllers/grade_controller.py - 成绩管理控制器

负责成绩录入、批量导入、成绩修改申请、成绩审核等业务逻辑。
所有写操作使用数据库事务控制，并写入操作日志。
"""

from datetime import datetime
from openpyxl import load_workbook

from backend.models.base import DatabaseManager
from backend.models.grade import Grade
from backend.models.enrollment import Enrollment
from backend.models.course_plan import CoursePlan
from backend.models.course import Course
from backend.models.student import Student
from backend.models.operation_log import OperationLog
from backend.utils.gpa_calculator import calculate_gpa
from backend.utils.validator import validate_score
from backend.utils.log_util import get_logger

logger = get_logger("grade_controller")


class GradeController:
    """成绩管理控制器。

    提供成绩录入、批量导入、成绩修改申请、成绩审核等功能。
    所有写操作均使用事务控制并记录审计日志。
    """

    def __init__(self):
        self._db = DatabaseManager.get_instance()

    def record_grade(self, teacher_id: str, student_id: str,
                     plan_id: int, score: int) -> dict:
        """录入单个学生的成绩。

        校验逻辑:
        1. 验证 score 是否为 0-100 之间的整数
        2. 查询 enrollment 表确认该学生已选此课
        3. 检查 grade 表是否已存在该记录（已存在提示走修改流程）
        4. 调用 gpa_calculator.calculate_gpa() 计算绩点
        5. 事务中写入 grade 表并记录日志

        Args:
            teacher_id: 教师工号。
            student_id: 学生学号。
            plan_id: 开课计划ID。
            score: 百分制成绩（0-100整数）。

        Returns:
            dict: {'success': bool, 'message': str}
        """
        try:
            # 校验成绩值
            valid, err = validate_score(score)
            if not valid:
                return {"success": False, "message": f"成绩格式错误: {err}"}

            score_int = int(score)

            with self._db.get_session() as session:
                # 确认学生已选此课
                enrollment = session.query(Enrollment).filter_by(
                    student_id=student_id,
                    plan_id=plan_id,
                    status="已选",
                ).first()
                if enrollment is None:
                    return {"success": False,
                            "message": "该学生未选此课程，无法录入成绩"}

                # 检查是否已存在成绩记录
                existing = session.query(Grade).filter_by(
                    student_id=student_id, plan_id=plan_id
                ).first()
                if existing:
                    return {
                        "success": False,
                        "message": (
                            "该学生成绩已存在，如需要修改请走修改审核流程"
                        ),
                    }

                # 计算绩点
                gpa_point = calculate_gpa(score_int)

                # 写入成绩
                grade = Grade(
                    student_id=student_id,
                    plan_id=plan_id,
                    score=score_int,
                    gpa_point=gpa_point,
                    record_time=datetime.now(),
                    status="正常",
                )
                session.add(grade)

                self._write_log(session, teacher_id, "成绩",
                                f"录入成绩: student={student_id}, "
                                f"plan={plan_id}, score={score_int}",
                                "成功", "")
                logger.info(
                    f"教师{teacher_id}录入学生{student_id}成绩: "
                    f"plan={plan_id}, score={score_int}, gpa={gpa_point}"
                )

                return {"success": True, "message": "成绩录入成功"}

        except Exception as e:
            logger.error(f"成绩录入异常: {e}", exc_info=True)
            return {"success": False, "message": "操作失败，请重试"}

    def batch_record_grade(self, teacher_id: str, plan_id: int,
                           file_path: str) -> dict:
        """批量导入成绩（通过Excel文件）。

        使用 openpyxl 解析 Excel 文件，第一行为表头，
        列格式: 学号、成绩。

        逐行校验，校验失败记录至 fail_list，
        批量 INSERT 通过的记录。

        Args:
            teacher_id: 教师工号。
            plan_id: 开课计划ID。
            file_path: Excel文件路径。

        Returns:
            dict: {'success_count': int, 'fail_count': int, 'fail_list': list}
        """
        fail_list = []
        success_list = []

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
        except Exception as e:
            logger.error(f"Excel文件读取失败: {e}")
            return {
                "success_count": 0,
                "fail_count": 0,
                "fail_list": [{"row": 0, "student_id": "", "reason": f"文件读取失败: {e}"}],
            }

        # 逐行校验
        for row_idx, row in enumerate(rows, 2):
            if not row or len(row) < 2:
                fail_list.append({
                    "row": row_idx,
                    "student_id": "",
                    "reason": "数据列不足",
                })
                continue

            student_id = str(row[0]).strip() if row[0] else ""
            score = row[1]

            # 学号非空
            if not student_id:
                fail_list.append({
                    "row": row_idx,
                    "student_id": "",
                    "reason": "学号为空",
                })
                continue

            # 成绩校验
            valid, err = validate_score(score)
            if not valid:
                fail_list.append({
                    "row": row_idx,
                    "student_id": student_id,
                    "reason": f"成绩非法: {err}",
                })
                continue

            success_list.append({
                "student_id": student_id,
                "score": int(score),
            })

        # 批量写入
        success_count = 0
        try:
            with self._db.get_session() as session:
                for item in success_list:
                    try:
                        # 确认选课记录
                        enrollment = session.query(Enrollment).filter_by(
                            student_id=item["student_id"],
                            plan_id=plan_id,
                            status="已选",
                        ).first()
                        if enrollment is None:
                            fail_list.append({
                                "row": 0,
                                "student_id": item["student_id"],
                                "reason": "未选此课",
                            })
                            continue

                        # 检查是否已有成绩
                        existing = session.query(Grade).filter_by(
                            student_id=item["student_id"],
                            plan_id=plan_id,
                        ).first()
                        if existing:
                            fail_list.append({
                                "row": 0,
                                "student_id": item["student_id"],
                                "reason": "成绩已存在",
                            })
                            continue

                        gpa_point = calculate_gpa(item["score"])
                        grade = Grade(
                            student_id=item["student_id"],
                            plan_id=plan_id,
                            score=item["score"],
                            gpa_point=gpa_point,
                            record_time=datetime.now(),
                            status="正常",
                        )
                        session.add(grade)
                        success_count += 1
                    except Exception as e:
                        fail_list.append({
                            "row": 0,
                            "student_id": item.get("student_id", ""),
                            "reason": f"写入异常: {e}",
                        })

                self._write_log(
                    session, teacher_id, "成绩",
                    f"批量导入: plan={plan_id}, 成功{success_count}条, "
                    f"失败{len(fail_list)}条",
                    "成功", ""
                )

        except Exception as e:
            logger.error(f"批量导入异常: {e}", exc_info=True)
            return {
                "success_count": 0,
                "fail_count": len(fail_list),
                "fail_list": fail_list,
            }

        result = {
            "success_count": success_count,
            "fail_count": len(fail_list),
            "fail_list": fail_list,
        }
        logger.info(
            f"批量导入完成: plan={plan_id}, "
            f"成功{success_count}, 失败{len(fail_list)}"
        )
        return result

    def apply_grade_modify(self, teacher_id: str, grade_id: int,
                           new_score: int, reason: str) -> dict:
        """申请成绩修改（教师提交修改请求）。

        验证成绩记录存在且属于该教师的课程，
        将成绩状态更新为"待审核"并记录修改原因。

        Args:
            teacher_id: 教师工号。
            grade_id: 成绩记录ID。
            new_score: 新成绩值。
            reason: 修改原因。

        Returns:
            dict: {'success': bool, 'message': str}
        """
        try:
            # 校验成绩值
            valid, err = validate_score(new_score)
            if not valid:
                return {"success": False, "message": f"成绩格式错误: {err}"}

            # reason非空校验
            if not reason or not reason.strip():
                return {"success": False, "message": "修改原因不能为空"}

            with self._db.get_session() as session:
                grade = session.query(Grade).filter_by(
                    grade_id=grade_id
                ).first()
                if grade is None:
                    return {"success": False,
                            "message": "成绩记录不存在"}

                # 验证成绩属于该教师的课程
                plan = session.query(CoursePlan).filter_by(
                    plan_id=grade.plan_id
                ).first()
                if plan is None or plan.teacher_id != teacher_id:
                    return {"success": False,
                            "message": "无权修改此成绩记录"}

                if grade.status != "正常":
                    return {"success": False,
                            "message": f"当前状态为'{grade.status}'，不可申请修改"}

                # 更新为待审核状态，记录新成绩和原因
                grade.status = "待审核"
                grade.modify_reason = f"申请修改为{new_score}: {reason.strip()}"

                self._write_log(
                    session, teacher_id, "成绩",
                    f"申请修改成绩: grade_id={grade_id}, "
                    f"{grade.score}→{new_score}, 原因: {reason}",
                    "成功", ""
                )

                logger.info(
                    f"教师{teacher_id}申请修改成绩grade_id={grade_id}"
                )
                return {"success": True, "message": "修改申请已提交，等待审核"}

        except Exception as e:
            logger.error(f"成绩修改申请异常: {e}", exc_info=True)
            return {"success": False, "message": "操作失败，请重试"}

    def audit_grade(self, admin_id: str, grade_id: int,
                    action: str, comment: str = "") -> dict:
        """审核成绩修改申请。

        action='approve': 通过申请，更新成绩和绩点，状态改为'已更正'
        action='reject': 驳回申请，状态恢复为'正常'，成绩不变

        Args:
            admin_id: 管理员账号。
            grade_id: 成绩记录ID。
            action: 审核操作，'approve'或'reject'。
            comment: 审核意见。

        Returns:
            dict: {'success': bool, 'message': str}
        """
        try:
            if action not in ("approve", "reject"):
                return {"success": False,
                        "message": "无效的审核操作"}

            with self._db.get_session() as session:
                grade = session.query(Grade).filter_by(
                    grade_id=grade_id
                ).first()
                if grade is None:
                    return {"success": False,
                            "message": "成绩记录不存在"}

                if grade.status != "待审核":
                    return {"success": False,
                            "message": f"当前状态为'{grade.status}'，无需审核"}

                if action == "approve":
                    # 从 modify_reason 中提取新成绩值
                    new_score = self._extract_new_score(
                        grade.modify_reason
                    )
                    if new_score is None:
                        return {"success": False,
                                "message": "无法获取修改后的成绩值"}

                    old_score = grade.score
                    grade.score = new_score
                    grade.gpa_point = calculate_gpa(new_score)
                    grade.status = "已更正"
                    if comment:
                        grade.modify_reason = (
                            (grade.modify_reason or "")
                            + f" [审核通过: {comment}]"
                        )
                    operation_desc = (
                        f"审核通过: grade_id={grade_id}, "
                        f"{old_score}→{new_score}"
                    )

                else:  # reject
                    grade.status = "正常"
                    if comment:
                        grade.modify_reason = (
                            (grade.modify_reason or "")
                            + f" [审核驳回: {comment}]"
                        )
                    operation_desc = (
                        f"审核驳回: grade_id={grade_id}"
                    )

                self._write_log(
                    session, admin_id, "成绩",
                    operation_desc, "成功", ""
                )

                logger.info(
                    f"管理员{admin_id}{'通过' if action == 'approve' else '驳回'}"
                    f"成绩审核: grade_id={grade_id}"
                )
                return {"success": True,
                        "message": f"审核{'通过' if action == 'approve' else '驳回'}"}

        except Exception as e:
            logger.error(f"成绩审核异常: {e}", exc_info=True)
            return {"success": False, "message": "操作失败，请重试"}

    def _extract_new_score(self, modify_reason: str) -> int:
        """从修改原因字符串中提取新成绩值。

        搜索模式: '申请修改为{score}' 或直接的数字。

        Args:
            modify_reason: 修改原因字符串。

        Returns:
            int: 新成绩值，解析失败返回None。
        """
        import re

        if not modify_reason:
            return None
        match = re.search(r'修改为(\d+)', modify_reason)
        if match:
            score = int(match.group(1))
            if 0 <= score <= 100:
                return score
        # 尝试查找任意数字
        numbers = re.findall(r'\d+', modify_reason)
        for num_str in numbers:
            score = int(num_str)
            if 0 <= score <= 100:
                return score
        return None

    def _write_log(self, session, user_id: str, log_type: str,
                   operation: str, result: str, ip_address: str = ""
                   ) -> None:
        """写入操作日志。"""
        try:
            log_entry = OperationLog(
                user_id=user_id,
                log_type=log_type,
                operation=operation,
                result=result,
                log_time=datetime.now(),
                ip_address=ip_address,
            )
            session.add(log_entry)
        except Exception as e:
            logger.warning(f"操作日志写入失败: {e}")
