"""
export_util.py - Excel/PDF导出工具

使用openpyxl生成Excel报表，支持表头加粗、数据行交替色等基本样式。
"""

import os
import zipfile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.utils.log_util import get_logger

logger = get_logger("export_util")

# 预定义样式
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2196F3", end_color="2196F3",
                          fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
EVEN_ROW_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD",
                            fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExportError(RuntimeError):
    """Raised when an export cannot be created or fails structural validation."""


def validate_xlsx(file_path, *, sheet_name, headers, merge_ranges=None):
    if not os.path.isfile(file_path) or os.path.getsize(file_path) < 1024:
        raise ExportError("导出文件不存在或大小异常")
    if not zipfile.is_zipfile(file_path):
        raise ExportError("导出文件不是有效的 XLSX 容器")
    try:
        workbook = load_workbook(file_path, read_only=False, data_only=False)
        if sheet_name not in workbook.sheetnames:
            raise ExportError("导出工作表缺失")
        sheet = workbook[sheet_name]
        actual_headers = [sheet.cell(1, i + 1).value for i in range(len(headers))]
        if actual_headers != headers:
            raise ExportError("导出表头校验失败")
        actual_merges = {str(item) for item in sheet.merged_cells.ranges}
        expected_merges = set(merge_ranges or [])
        if not expected_merges.issubset(actual_merges):
            raise ExportError("导出合并单元格校验失败")
        workbook.close()
    except ExportError:
        raise
    except Exception as exc:
        raise ExportError("导出文件无法重新打开") from exc
    return file_path


def export_to_excel(
    headers: list,
    data_rows: list,
    file_path: str,
    sheet_name: str = "Sheet1",
    summary_row: dict = None,
) -> str:
    """将数据导出为Excel文件。

    Args:
        headers: 表头列名列表，如 ['学号', '姓名', '成绩']。
        data_rows: 数据行列表，每项为list，与headers一一对应。
        file_path: 导出文件保存路径（含.xlsx扩展名）。
        sheet_name: 工作表名称。
        summary_row: 可选的汇总行，dict形式如 {'学号': '', '姓名': '合计', '成绩': 350}。

    Returns:
        bool: 导出成功返回True，失败返回False。
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # 写入数据行（交替行色）
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = EVEN_ROW_FILL

        # 写入汇总行
        if summary_row:
            summary_row_idx = len(data_rows) + 2
            for col_idx, header in enumerate(headers, 1):
                value = summary_row.get(header, "")
                cell = ws.cell(row=summary_row_idx, column=col_idx,
                               value=value)
                cell.font = Font(name="微软雅黑", size=11, bold=True)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            max_width = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        # 中文字符按2个字符宽度计算
                        cell_len = sum(
                            2 if ord(c) > 127 else 1 for c in str(cell.value)
                        )
                        max_width = max(max_width, cell_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                min(max_width + 4, 40)
            )

        # 确保目标目录存在
        os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path)
                    else ".", exist_ok=True)
        wb.save(file_path)
        logger.info(f"Excel导出成功: {file_path}")
        return validate_xlsx(
            file_path, sheet_name=sheet_name, headers=headers
        )
    except Exception as e:
        logger.error(f"Excel导出失败: {e}")
        raise ExportError("Excel 导出失败") from e


def export_schedule_to_excel(
    headers: list,
    data_rows: list,
    merge_ranges: list,
    file_path: str,
    sheet_name: str = "个人课表",
) -> str:
    """导出带合并单元格的课表Excel文件。

    Args:
        headers: 表头列名列表（含"节次"和"时间"两列）。
        data_rows: 数据行列表，每项为list。
        merge_ranges: openpyxl merge_cells 范围字符串列表，
                      如 ['C3:C5', 'D6:D8']。
        file_path: 导出文件保存路径。
        sheet_name: 工作表名称。

    Returns:
        bool: 导出成功返回True，失败返回False。
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = EVEN_ROW_FILL

        for merge_range in merge_ranges:
            ws.merge_cells(merge_range)

        for col_idx in range(1, len(headers) + 1):
            max_width = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_len = sum(
                            2 if ord(c) > 127 else 1 for c in str(cell.value)
                        )
                        max_width = max(max_width, cell_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                min(max_width + 4, 40)
            )

        os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path)
                    else ".", exist_ok=True)
        wb.save(file_path)
        logger.info(f"课表Excel导出成功: {file_path}")
        return validate_xlsx(
            file_path,
            sheet_name=sheet_name,
            headers=headers,
            merge_ranges=merge_ranges,
        )
    except Exception as e:
        logger.error(f"课表Excel导出失败: {e}")
        raise ExportError("课表 Excel 导出失败") from e
