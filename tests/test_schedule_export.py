"""Critical schedule-grid and XLSX placement regression tests."""

import os
import tempfile
import unittest

from openpyxl import load_workbook


class TestScheduleExport(unittest.TestCase):
    def test_two_period_course_starts_at_c2_and_merges_c2_c3(self):
        from backend.services.schedule_grid import build_schedule_grid
        from backend.utils.export_util import export_schedule_to_excel

        course = {
            "plan_id": 1,
            "course_id": "CS101",
            "course_name": "程序设计",
            "semester": "2026-2027-1",
            "weekday": 1,
            "period_start": 1,
            "period_count": 2,
            "start_week": 1,
            "end_week": 18,
            "location": "教一101",
        }
        model = build_schedule_grid([course], "2026-2027-1")
        self.assertIn("C2:C3", model["merge_ranges"])
        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "schedule.xlsx")
            headers = ["节次", "时间", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            export_schedule_to_excel(headers, model["schedule"], model["merge_ranges"], path)
            workbook = load_workbook(path)
            sheet = workbook["个人课表"]
            self.assertIn("程序设计", sheet["C2"].value)
            self.assertIn("C2:C3", {str(item) for item in sheet.merged_cells.ranges})
            self.assertIsNone(sheet["C3"].value)
            workbook.close()

    def test_different_spans_in_same_cell_remain_atomic(self):
        from backend.services.schedule_grid import build_schedule_grid

        base = {
            "semester": "2026-2027-1", "weekday": 1, "period_start": 1,
            "start_week": 1, "end_week": 18, "location": "",
        }
        courses = [
            {**base, "plan_id": 1, "course_name": "长课", "period_count": 2},
            {**base, "plan_id": 2, "course_name": "短课", "period_count": 1},
        ]
        model = build_schedule_grid(courses, "2026-2027-1")
        self.assertNotIn("C2:C3", model["merge_ranges"])
        self.assertIn("长课", model["schedule"][0][2])
        self.assertIn("短课", model["schedule"][0][2])
        self.assertIn("长课", model["schedule"][1][2])

    def test_general_export_is_reopened_and_validated(self):
        from backend.utils.export_util import export_to_excel

        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "grades.xlsx")
            result = export_to_excel(
                ["学号", "成绩"],
                [["S001", 90]],
                path,
                sheet_name="成绩",
                summary_row={"学号": "汇总", "成绩": 90},
            )
            self.assertEqual(result, path)
            workbook = load_workbook(path)
            self.assertEqual(workbook["成绩"]["A2"].value, "S001")
            workbook.close()

    def test_invalid_schedule_ranges_fail_closed(self):
        from backend.api.errors import ApiError
        from backend.services.schedule_grid import build_schedule_grid

        invalid = {
            "plan_id": 1, "course_name": "越界课程", "semester": "2026-2027-1",
            "weekday": 1, "period_start": 11, "period_count": 2,
            "start_week": 1, "end_week": 18,
        }
        with self.assertRaises(ApiError):
            build_schedule_grid([invalid], "2026-2027-1")

    def test_export_validator_rejects_bad_container_and_structure(self):
        from backend.utils.export_util import ExportError, validate_xlsx

        with tempfile.TemporaryDirectory() as directory:
            missing = os.path.join(directory, "missing.xlsx")
            with self.assertRaises(ExportError):
                validate_xlsx(missing, sheet_name="Sheet", headers=["A"])

            bad = os.path.join(directory, "bad.xlsx")
            with open(bad, "wb") as output:
                output.write(b"not an xlsx" * 200)
            with self.assertRaises(ExportError):
                validate_xlsx(bad, sheet_name="Sheet", headers=["A"])

            valid = os.path.join(directory, "valid.xlsx")
            from openpyxl import Workbook
            book = Workbook(); book.active.title = "Actual"; book.active["A1"] = "Wrong"; book.save(valid)
            with self.assertRaises(ExportError):
                validate_xlsx(valid, sheet_name="Missing", headers=["A"])
            with self.assertRaises(ExportError):
                validate_xlsx(valid, sheet_name="Actual", headers=["A"])
            with self.assertRaises(ExportError):
                validate_xlsx(valid, sheet_name="Actual", headers=["Wrong"], merge_ranges=["A1:A2"])


if __name__ == "__main__":
    unittest.main()
